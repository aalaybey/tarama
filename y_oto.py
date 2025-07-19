import sys
import os
import json
import requests
import shutil
import math
from datetime import datetime, timedelta, UTC
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
import re
import time
import threading
from io import BytesIO
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import portalocker
import psycopg2

DAYS = 100
AWS_BUCKET = "alaybey"
AWS_REGION = os.getenv("AWS_REGION")
AWS_ACCESS_KEY_ID = os.getenv("AWS_ACCESS")
AWS_SECRET_ACCESS_KEY = os.getenv("AWS_SECRET")
ENDPOINT_URL = f"https://s3.{AWS_REGION}.wasabisys.com"
USER_AGENT = "Alper Alaybey <a.alaybey@gmail.com>"
HEADERS = {"User-Agent": USER_AGENT}

import boto3
from botocore.client import Config
s3 = boto3.client(
    "s3",
    region_name=AWS_REGION,
    aws_access_key_id=AWS_ACCESS_KEY_ID,
    aws_secret_access_key=AWS_SECRET_ACCESS_KEY,
    endpoint_url=ENDPOINT_URL,
    config=Config(signature_version="s3v4"),
)

def s3_path(key):
    rel_path = key.replace("\\", "/").replace("./", "")
    return rel_path.lstrip("/")

def s3_exists(key):
    try:
        s3.head_object(Bucket=AWS_BUCKET, Key=s3_path(key))
        return True
    except:
        return False

def s3_read_text(key):
    obj = s3.get_object(Bucket=AWS_BUCKET, Key=s3_path(key))
    return obj["Body"].read().decode("utf-8")

def s3_read_bytes(key):
    obj = s3.get_object(Bucket=AWS_BUCKET, Key=s3_path(key))
    return obj["Body"].read()

def s3_write_text(key, text):
    s3.put_object(Bucket=AWS_BUCKET, Key=s3_path(key), Body=text.encode("utf-8"))

def s3_write_bytes(key, b):
    s3.put_object(Bucket=AWS_BUCKET, Key=s3_path(key), Body=b)

def s3_delete(key):
    s3.delete_object(Bucket=AWS_BUCKET, Key=s3_path(key))

def s3_list_dir(prefix):
    paginator = s3.get_paginator("list_objects_v2")
    result = paginator.paginate(Bucket=AWS_BUCKET, Prefix=s3_path(prefix))
    return [content['Key'] for page in result for content in page.get('Contents', [])]

def get_trigger_ticker():
    key = "trigger.txt"
    if not s3_exists(key):
        raise Exception("trigger.txt bulunamadı!")
    ticker = s3_read_text(key).strip()
    if not ticker:
        raise Exception("trigger.txt içinde ticker bulunamadı!")
    return ticker.upper(), key

def get_cik_for_ticker(ticker, tickers_file="tickers.txt"):
    with open(tickers_file, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or ',' not in line:
                continue
            t, cik = line.split(",", 1)
            if t.strip().upper() == ticker.upper():
                return cik.strip().zfill(10)
    raise Exception(f"{ticker} için cik bulunamadı (tickers.txt'de yok)")

consecutive_error_count = 0
REQUEST_COUNT = 0
REQUEST_LOCK = threading.Lock()

def incr_request_and_sleep():
    global REQUEST_COUNT
    with REQUEST_LOCK:
        REQUEST_COUNT += 1
        if REQUEST_COUNT % 10 == 0:
            print(f"🕒 {REQUEST_COUNT} request atıldı, 2 sn bekleniyor...")
            time.sleep(0.5)

def inc_error_and_kill_if_limit(limit=50):
    global consecutive_error_count
    consecutive_error_count += 1
    if consecutive_error_count >= limit:
        print("LOG: Script kill oldu! (üst üste hata limiti aşıldı)")
        print(f"❌ ÜST ÜSTE {limit} HATA! Script kill ediliyor.")
        sys.exit(1)

def download_file(url, s3key):
    backoff_count = 0
    while True:
        try:
            incr_request_and_sleep()
            resp = requests.get(url, headers=HEADERS)
            if resp.status_code in [429, 403]:
                backoff_count += 1
                print(f"⏳ Rate-limit algılandı! {backoff_count}. kez 2 dakika bekleniyor... [download_file] {url}")
                if backoff_count >= 3:
                    print("❌ 3 kez üst üste backoff, script kill ediliyor!")
                    sys.exit(1)
                time.sleep(120)
                continue
            if resp.status_code >= 400:
                print(f"📥 Dosya indirme hatası: {url} => {resp.status_code} {resp.reason}")
                break
            s3_write_bytes(s3key, resp.content)
            break
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout):
            backoff_count += 1
            print(f"🌐 Bağlantı hatası: {url}")
            if backoff_count >= 3:
                print("❌ 3 kez üst üste backoff, script kill ediliyor!")
                sys.exit(1)
            time.sleep(120)
            continue
        except Exception as e:
            print(f"📥 Dosya indirme hatası: {url} => {e}")
            inc_error_and_kill_if_limit()
            break

def clean_hidden_rows_from_html(html):
    soup = BeautifulSoup(html, 'lxml')
    for tr in soup.find_all('tr', style=lambda x: x and 'display:none' in x):
        tr.decompose()
    for td in soup.find_all('td', style=lambda x: x and 'display:none' in x):
        td.decompose()
    for th in soup.find_all('th', style=lambda x: x and 'display:none' in x):
        th.decompose()
    return str(soup)

num_pat = re.compile(r'[^0-9().-]')

def to_number(x):
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return None if (math.isnan(x) or math.isinf(x)) else float(x)
    s = num_pat.sub('', str(x))
    if not s or s in ('-', '.', '()'):
        return None
    neg = s.startswith('(') and s.endswith(')')
    if neg:
        s = s[1:-1]
    try:
        v = float(s)
        if math.isnan(v) or math.isinf(v):
            return None
        return -v if neg else v
    except ValueError:
        return None

def get_sec_html_url(cik, accession):
    return f"https://www.sec.gov/cgi-bin/viewer?action=view&cik={cik}&accession_number={accession}&xbrl_type=v#"

def get_quarter_from_date(report_date, form_type):
    year = str(report_date.year)
    month = report_date.month
    quarter_idx = (month - 1) // 3 + 1
    quarter = f"Q{quarter_idx}"
    if form_type == "10-K":
        quarter = "Q4"
    return year, quarter

def get_quarter_from_cover_page(driver, default_year, default_quarter, form_type):
    html = driver.page_source
    year = default_year
    quarter = default_quarter
    m = re.search(r'Document\s*Fiscal\s*Year\s*Focus.*?([12][0-9]{3})', html, flags=re.I | re.S)
    if m:
        year = m.group(1)
    pm = re.search(r'Document\s*Fiscal\s*Period\s*Focus.*?(Q[1-4]|0?[1-4]|FY)', html, flags=re.I | re.S)
    if pm:
        token = pm.group(1).upper()
        if token.startswith('Q'):
            quarter = token
        elif token == 'FY':
            quarter = 'Q4'
        else:
            quarter = f"Q{int(token)}"
    if form_type == "10-K":
        quarter = "Q4"
    return str(year), quarter

def detect_period_col_from_row2(df, quarter):
    if df is None or df.empty or not quarter:
        return None
    try:
        row2 = df.iloc[2]
    except Exception:
        return None
    quarter_map = {
        "Q1": "3 month",
        "Q2": "6 month",
        "Q3": "9 month",
        "Q4": "12 month",
    }
    target = quarter_map.get(quarter.upper())
    if not target:
        return None
    target = target.lower()
    for idx, val in enumerate(row2):
        if pd.isna(val):
            continue
        if target in str(val).lower():
            return idx
    return None

def detect_share_multiplier_from_text(title):
    t = title.lower()
    if "shares in millions" in t:
        return 1_000_000.0
    elif "shares in thousands" in t:
        return 1_000.0
    return 1.0

def detect_multiplier_from_text(title):
    t = title.lower()
    if "millions" in t:
        return 1_000_000.0
    elif "thousands" in t:
        return 1_000.0
    return 1.0

cash_flow_metrics = {
    "Depreciation and Amortization": [["depreciation", "amortization"]],
    "Amortization": [["amortization"]],
    "Depreciation": [["depreciation"]],
    "Cash From Operations": [["cash", "operat"]],
    "PPE Purchase": [["property", "equipment", "purchas"], ["property", "equipment", "add"], ["property", "equipment", "payment"], ["property", "equipment", "acqui"], ["property", "add"], ["property", "purchas"], ["property", "acqui"], ["capital", "expend"]],
    "PPE Sale": [["sale", "property"], ["sale", "fixed"], ["disposal", "property"], ["disposal", "fixed"], ["sale", "tangible"], ["disposal", "tangible"]],
    "Cash Business Acquisitions": [["business"]],
    "Dividends": [["payment", "dividend"]],
    "Cash Taxes": [["paid", "tax"], ["cash", "tax"], ["tax"]],
    "Cash Interest": [["paid", "interest"], ["cash", "interest"], ["interest"]],
}

income_metrics = {
    "Sales": [["revenue"], ["sale"]],
    "Cost of Sales": [["cost of"]],
    "Gross Profit": [["gross"], ["margin"], ["gross", "profit"]],
    "EBIT": [["operating", "income"], ["operating", "loss"], ["income", "operations"], ["loss", "operations"]],
    "Net Income": [["net income"], ["net loss"], ["net", "income"], ["net", "loss"]],
    "Interest Income": [["interest", "income"], ["other", "net"]],
    "Interest Expense": [["interest", "expense"]],
    "EPS": [["basic", "net", "income"], ["basic", "net", "loss"], ["basic", "dollar"], ["basic", "usd"]],
    "Shares Outstanding": [["basic"]],
}

balance_metrics = {
    "Total Assets": [["total assets"]],
    "Total Equity": [["total", "equity"]],
    "Noncontrolling interest": [["non", "control", "interest"]],
    "Cash and cash equivalents": [["cash", "equivalents"]],
    "Inventories": [["inventor"]],
    "Accounts receivable": [["accounts", "receivable"]],
    "Prepaid expenses": [["prepaid", "expenses"]],
    "PPE": [["property", "equipment"]],
    "Intangible assets": [["intangible", "asset"]],
    "Operating lease assets": [["leas", "operati", "asset"]],
    "Digital assets": [["digital", "asset"]],
    "Right-of-use assets": [["use", "asset"]],
    "Accounts payable": [["account", "payable"]],
    "Operating lease liabilities": [["leas", "operati", "liab"]],
    "Accrued liabilities": [["accrue", "liabilit"], ["accrue", "expense"]],
    "Deferred revenue": [["defer", "revenue"], ["unearn", "revenue"]]
}

def find_metric_in_df(
    df, kw_groups, multiplier=1.0,
    exclude_term=None, reverse=False,
    quarter=None, row_start=None, col_override=None,
    tab_type=None,
    metric_name=None,
    ebit_row_idx=None,
    html_raw=None
):
    from bs4 import BeautifulSoup

    if metric_name in ("Cash Interest", "Cash Taxes") and html_raw is not None:
        soup = BeautifulSoup(html_raw, "lxml")
        all_tr = []
        for tr in soup.find_all("tr"):
            style = tr.get("style", "")
            if "display:none" in style:
                continue
            row_texts = [td.get_text(separator=" ", strip=True).lower() for td in tr.find_all(["td", "th"])]
            if row_texts:
                all_tr.append(row_texts)
        supplemental_idx = None
        for i, row in enumerate(all_tr[3:], start=3):
            if any("supplemental" in x for x in row):
                supplemental_idx = i
                break
        if supplemental_idx is None:
            return 0.0
        for row in all_tr[supplemental_idx+1:]:
            row_join = " ".join(row)
            if exclude_term and exclude_term.lower() in row_join:
                continue
            for group in kw_groups:
                for idx, cell in enumerate(row):
                    if all(k.lower() in cell for k in group):
                        for val in row[idx+1:]:
                            num = to_number(val)
                            if num is not None:
                                return num * multiplier
        return 0.0

    if metric_name == "Noncontrolling interest":
        rows_all = df.values.tolist()[::-1]
        for group in kw_groups:
            for row in rows_all:
                row_text = " ".join(str(x).lower() for x in row if pd.notna(x))
                if exclude_term and exclude_term.lower() in row_text:
                    continue
                idx_match = None
                for idx, cell in enumerate(row):
                    cell_str = str(cell).lower()
                    if all(k.lower() in cell_str for k in group):
                        idx_match = idx
                        break
                if idx_match is not None:
                    for c in row[idx_match+1:]:
                        v = to_number(c)
                        if v is not None:
                            return v * multiplier
                    for c in row:
                        v = to_number(c)
                        if v is not None:
                            return v * multiplier
        return 0.0

    if metric_name == "Right-of-use assets":
        rows_all = df.values.tolist()
        for group in kw_groups:
            for row in rows_all:
                row_text = " ".join(str(x).lower() for x in row if pd.notna(x))
                if "lease" in row_text:
                    continue
                if exclude_term and exclude_term.lower() in row_text:
                    continue
                idx_match = None
                for idx, cell in enumerate(row):
                    cell_str = str(cell).lower()
                    if all(k.lower() in cell_str for k in group):
                        idx_match = idx
                        break
                if idx_match is not None:
                    for c in row[idx_match+1:]:
                        v = to_number(c)
                        if v is not None:
                            return v * multiplier
                    for c in row:
                        v = to_number(c)
                        if v is not None:
                            return v * multiplier
        return 0.0

    if metric_name in ("Interest Income", "Interest Expense") and ebit_row_idx is not None:
        rows_all = df.values.tolist()[ebit_row_idx + 1:] if ebit_row_idx + 1 < len(df.values.tolist()) else []
        if not rows_all:
            return 0.0
    else:
        rows_all = df.values.tolist()

    rows = rows_all[::-1] if reverse else rows_all
    if row_start is not None:
        rows = rows[row_start:]

    quarter_map = {
        "Q1": ["3 month", "three months"],
        "Q2": ["6 month", "six months"],
        "Q3": ["9 month", "nine months"],
        "Q4": ["12 month", "twelve months", "year", "annual"],
    }
    quarter_tokens = quarter_map.get((quarter or "").upper(), [])
    quarter_col_idx = col_override
    if quarter_col_idx is None and quarter and quarter_tokens:
        for row in rows:
            for idx, cell in enumerate(row):
                val = str(cell).lower()
                for token in quarter_tokens:
                    if token in val:
                        quarter_col_idx = idx
                        break
                if quarter_col_idx is not None:
                    break
            if quarter_col_idx is not None:
                break

    for group in kw_groups:
        for row in rows:
            row_text = " ".join(str(x).lower() for x in row if pd.notna(x))
            if exclude_term and exclude_term.lower() in row_text:
                continue
            idx_match = None
            for idx, cell in enumerate(row):
                cell_str = str(cell).lower()
                if all(k.lower() in cell_str for k in group):
                    idx_match = idx
                    break
            if idx_match is not None:
                if tab_type == "income_statement":
                    q = (quarter or "").upper()
                    if q in ["Q2", "Q3"] and len(row) > idx_match + 3:
                        v = to_number(row[idx_match + 3])
                        if v is not None:
                            return v * multiplier
                    elif len(row) > idx_match + 1:
                        v = to_number(row[idx_match + 1])
                        if v is not None:
                            return v * multiplier
                if quarter_col_idx is not None and len(row) > quarter_col_idx:
                    v = to_number(row[quarter_col_idx])
                    if v is not None:
                        return v * multiplier
                for c in row[idx_match+1:]:
                    v = to_number(c)
                    if v is not None:
                        return v * multiplier
                for c in row:
                    v = to_number(c)
                    if v is not None:
                        return v * multiplier
    return None

def extract_tabular_data_from_html(driver):
    import time
    from selenium.webdriver.common.by import By
    print("="*60)
    print("extract_tabular_data_from_html ÇAĞRILDI")
    time.sleep(2)
    found_tables = []
    fs_menu = None
    try:
        fs_menu = driver.find_element(By.ID, "menu_cat3")
        print("[LOG] Menü bulundu: id=menu_cat3")
    except Exception as e:
        print("[LOG] Menü id=menu_cat3 bulunamadı:", e)
    if fs_menu is None:
        try:
            for a in driver.find_elements(By.CSS_SELECTOR, 'a[id^="menu_cat"]'):
                txt = (a.text or "").lower()
                print(f"[LOG] Alternatif menü adayı: {txt}")
                if "financial" in txt and "statement" in txt:
                    fs_menu = a
                    print("[LOG] Menü bulundu (alternatif yolla):", txt)
                    break
        except Exception as e:
            print("[LOG] Alternatif menü bulma hatası:", e)
    if fs_menu:
        try:
            driver.execute_script("arguments[0].scrollIntoView();", fs_menu)
            for _ in range(2):
                try:
                    fs_menu.click()
                    print("[LOG] Menüye tıklandı.")
                    time.sleep(2)
                    break
                except Exception as e:
                    print("[LOG] Menüye tıklanamadı, tekrar deneniyor:", e)
                    time.sleep(2)
        except Exception as e:
            print("[LOG] Menü scroll/tıkla hatası:", e)
    else:
        print("[LOG] Menü bulunamadı, yedek plan devreye alınacak.")
    links = []
    if fs_menu:
        try:
            ul = fs_menu.find_element(By.XPATH, 'following-sibling::ul')
            links = ul.find_elements(By.CSS_SELECTOR, 'a.xbrlviewer')
            print(f"[LOG] Menü altında {len(links)} adet tablo linki bulundu.")
        except Exception as e:
            print("[LOG] Menü altında tablo linki bulunamadı:", e)
    if not links:
        links = driver.find_elements(By.CSS_SELECTOR, 'a.xbrlviewer')
        print(f"[LOG] Yedek planda {len(links)} adet tablo linki bulundu.")
    print("\nTABLO LİSTESİ:")
    for link in links:
        print(" >", (link.text or link.get_attribute("innerText") or "").strip())
    print("\nBAŞLIK EŞLEŞME LOGU:")
    for link in links:
        link_text = (link.text or link.get_attribute("innerText") or "").strip()
        t = link_text.lower()
        matched = False
        if "balance sheet" in t and "parenthetical" not in t:
            print(f"BALANCE eşleşti  ---> {t}")
            found_tables.append(("balance_sheet", link_text, link))
            matched = True
        elif "cash flow" in t:
            print(f"CASH FLOW eşleşti  ---> {t}")
            found_tables.append(("cash_flow", link_text, link))
            matched = True
        elif ("income" in t and "statement" in t) or "statements of operations" in t:
            print(f"INCOME eşleşti  ---> {t}")
            found_tables.append(("income_statement", link_text, link))
            matched = True
        if not matched:
            print(f"EŞLEŞME YOK    ---> {t}")
    print("\n[LOG] found_tables:", [(a, b) for a, b, c in found_tables])
    print("="*60)
    return found_tables

def extract_metrics_from_sec_html(driver, found_tables, year, quarter):
    extracted = {}
    processed_types = set()
    dfs_in_ram = []
    for tab_type, link_text, link in found_tables:
        if tab_type in processed_types:
            continue
        try:
            link.click()
            import time
            time.sleep(2)
            html = driver.page_source
            clean_html = clean_hidden_rows_from_html(html)
            dfs = pd.read_html(clean_html)
            if not dfs:
                continue
            df = dfs[0]
            forced_col_idx = detect_period_col_from_row2(df, quarter)
            title = link_text.lower()
            try:
                row2 = " ".join([str(x).lower() for x in df.iloc[1] if pd.notna(x)])
            except Exception:
                row2 = ""
            if "$ in millions" in row2:
                multiplier = 1_000_000.0
            elif "$ in thousands" in row2:
                multiplier = 1_000.0
            else:
                multiplier = 1.0
            shares_out_multiplier = 1.0
            if tab_type == "income_statement":
                if "shares in millions" in row2:
                    shares_out_multiplier = 1_000_000.0
                elif "shares in thousands" in row2:
                    shares_out_multiplier = 1_000.0
                else:
                    shares_out_multiplier = 1.0
            if tab_type == "cash_flow":
                for metric, kw_groups in cash_flow_metrics.items():
                    val = find_metric_in_df(
                        df, kw_groups, multiplier=multiplier,
                        quarter=quarter, col_override=forced_col_idx,
                        tab_type=tab_type,
                        metric_name=metric,
                        html_raw=clean_html
                    )
                    extracted[metric] = val if val is not None else 0.0
                processed_types.add("cash_flow")
            elif tab_type == "income_statement":
                ebit_row_idx = None
                rows = df.values.tolist()
                for i, group in enumerate(income_metrics["EBIT"]):
                    for row_idx, row in enumerate(rows):
                        for cell in row:
                            cell_str = str(cell).lower()
                            if all(k.lower() in cell_str for k in group):
                                ebit_row_idx = row_idx
                                break
                        if ebit_row_idx is not None:
                            break
                    if ebit_row_idx is not None:
                        break
                for metric, kw_groups in income_metrics.items():
                    if metric == "EPS":
                        val = find_metric_in_df(
                            df, kw_groups, multiplier=1.0,
                            quarter=quarter, col_override=forced_col_idx,
                            tab_type=tab_type,
                            metric_name=metric,
                            html_raw=clean_html
                        )
                        extracted[metric] = val if val is not None else 0.0
                    elif metric == "Shares Outstanding":
                        val = find_metric_in_df(
                            df, kw_groups, multiplier=shares_out_multiplier, reverse=True,
                            quarter=quarter, col_override=forced_col_idx,
                            tab_type=tab_type,
                            metric_name=metric,
                            html_raw=clean_html
                        )
                        extracted[metric] = val if val is not None else 0.0
                    elif metric in ("Interest Income", "Interest Expense"):
                        val = find_metric_in_df(
                            df, kw_groups, multiplier=multiplier,
                            quarter=quarter, col_override=forced_col_idx,
                            tab_type=tab_type,
                            metric_name=metric,
                            ebit_row_idx=ebit_row_idx,
                            html_raw=clean_html
                        )
                        extracted[metric] = val if val is not None else 0.0
                    else:
                        val = find_metric_in_df(
                            df, kw_groups, multiplier=multiplier,
                            quarter=quarter, col_override=forced_col_idx,
                            tab_type=tab_type,
                            metric_name=metric,
                            html_raw=clean_html
                        )
                        extracted[metric] = val if val is not None else 0.0
                processed_types.add("income_statement")
            elif tab_type == "balance_sheet":
                for metric, kw_groups in balance_metrics.items():
                    val = find_metric_in_df(
                        df, kw_groups, multiplier=multiplier,
                        quarter=quarter, col_override=forced_col_idx,
                        tab_type=tab_type,
                        metric_name=metric,
                        html_raw=clean_html
                    )
                    extracted[metric] = val if val is not None else 0.0
                processed_types.add("balance_sheet")
            del df
            del dfs
        except Exception as e:
            print("[ERROR] extract_metrics_from_sec_html:", str(e))
            continue
    for df in dfs_in_ram:
        del df
    return extracted

def sort_quarter_columns(cols):
    def quarter_key(col):
        import re
        m = re.match(r"(\d{4})\s+Q([1-4])", str(col))
        if not m:
            return (0, 0)
        year = int(m.group(1))
        quarter = int(m.group(2))
        return (-year, -quarter)
    if "Metric" in cols:
        metric_idx = cols.index("Metric")
        main_cols = [c for i, c in enumerate(cols) if i != metric_idx]
        sorted_main = sorted(main_cols, key=quarter_key)
        return ["Metric"] + sorted_main
    else:
        return sorted(cols, key=quarter_key)

def save_to_final_excel(symbol, year, quarter, extracted):
    import os
    os.makedirs("Final", exist_ok=True)
    out_path = os.path.join("Final", f"{symbol}.xlsx")
    all_metrics = list(cash_flow_metrics.keys()) + list(income_metrics.keys()) + list(balance_metrics.keys())
    col_name = f"{year} {quarter}"
    col_values = [extracted.get(m, 0.0) for m in all_metrics]
    if os.path.exists(out_path):
        try:
            df_existing = pd.read_excel(out_path)
            if "Metric" not in df_existing.columns:
                df_existing.insert(0, "Metric", all_metrics)
            df = df_existing.set_index("Metric")
        except Exception:
            df = pd.DataFrame(index=all_metrics)
    else:
        df = pd.DataFrame(index=all_metrics)
    df[col_name] = col_values
    df_reset = df.reset_index()
    cols = df_reset.columns.tolist()
    sorted_cols = sort_quarter_columns(cols)
    df_sorted = df_reset[sorted_cols]
    df_sorted.to_excel(out_path, index=False)
    try:
        wb_out = openpyxl.load_workbook(out_path)
        ws = wb_out.active
        for row in ws.iter_rows(min_row=2, min_col=2):
            for cell in row:
                if cell.value is None:
                    cell.value = ""
        wb_out.save(out_path)
    except Exception:
        pass

def main():
    ticker, trigger_key = get_trigger_ticker()
    cik = get_cik_for_ticker(ticker)
    days = DAYS
    url = f"https://data.sec.gov/submissions/CIK{cik}.json"
    backoff_count = 0
    while True:
        try:
            incr_request_and_sleep()
            r = requests.get(url, headers=HEADERS)
            if r.status_code in [429, 403]:
                backoff_count += 1
                print(f"⏳ Rate-limit algılandı! {backoff_count}. kez 2 dakika bekleniyor... [main] {url}")
                if backoff_count >= 3:
                    print("❌ 3 kez üst üste backoff, script kill ediliyor!")
                    sys.exit(1)
                time.sleep(120)
                continue
            if r.status_code >= 400:
                print(f"⚠️ {ticker}: {cik} - API hatası veya veri yok. {r.status_code} {r.reason}")
                inc_error_and_kill_if_limit()
                return
            data = r.json().get("filings", {}).get("recent", {})
            break
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout):
            backoff_count += 1
            print(f"🌐 Bağlantı hatası: {url}")
            if backoff_count >= 3:
                print("❌ 3 kez üst üste backoff, script kill ediliyor!")
                sys.exit(1)
            time.sleep(120)
            continue
        except Exception as e:
            print(f"⚠️ {ticker}: {cik} - Veri çekilemedi: {e}")
            inc_error_and_kill_if_limit()
            return
    forms = data.get("form", [])
    accessions = data.get("accessionNumber", [])
    report_dates = data.get("reportDate", [])
    found = False
    for i, ftype in enumerate(forms):
        if ftype not in ["10-Q", "10-K"]:
            continue
        try:
            rpt_date = datetime.strptime(report_dates[i], "%Y-%m-%d")
        except:
            continue
        filter_date = (datetime.now(UTC) - timedelta(days=days)).date()
        if rpt_date.date() < filter_date:
            continue
        acc = accessions[i].replace("-", "")
        year, quarter = get_quarter_from_date(rpt_date, ftype)
        html_url = get_sec_html_url(cik, acc)
        chrome_options = Options()
        chrome_options.add_argument("--headless=new")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--window-size=1920,1080")
        chrome_options.add_argument(f"user-agent={USER_AGENT}")
        driver = webdriver.Chrome(options=chrome_options)
        driver.get(html_url)
        time.sleep(2)
        year, quarter = get_quarter_from_cover_page(driver, year, quarter, ftype)
        found_tables = extract_tabular_data_from_html(driver)
        if found_tables:
            extracted = extract_metrics_from_sec_html(driver, found_tables, year, quarter)
            save_to_final_excel(ticker, year, quarter, extracted)
            found = True
        driver.quit()
        if found:
            break
    if not found:
        print(f"{ticker} için uygun SEC verisi bulunamadı.")

import yfinance as yf

def fill_dates_and_prices_in_ws(ws_dst):
    ticker = ws_dst["B40"].value
    index_ticker = ws_dst["C40"].value
    if not ticker:
        return
    if not index_ticker:
        return
    if index_ticker.upper().startswith("ARCX:"):
        index_ticker = index_ticker.split(":", 1)[1]
    index_ticker = index_ticker.strip()
    start_row = 41
    end_row = 107
    curr_date = datetime.today().date()
    min_date = curr_date - timedelta(days=(end_row - start_row) * 30 + 30)
    def get_hist(ticker_code):
        try:
            yf_ticker = yf.Ticker(ticker_code)
            hist = yf_ticker.history(start=min_date.strftime("%Y-%m-%d"), end=(curr_date + timedelta(days=1)).strftime("%Y-%m-%d"))
            hist = hist.reset_index()
            if 'Date' in hist:
                hist['Date'] = hist['Date'].dt.date
            return hist
        except Exception as e:
            return None
    hist_company = get_hist(ticker)
    hist_index = get_hist(index_ticker)
    prev_date = curr_date
    for i, row in enumerate(range(start_row, end_row + 1)):
        if i == 0:
            this_date = curr_date
        else:
            this_date = prev_date - timedelta(days=30)
        def find_price(hist, date):
            search_date = date
            close_price = None
            while close_price is None and search_date >= min_date:
                if hist is not None:
                    row_df = hist[hist['Date'] == search_date]
                    if not row_df.empty and not row_df['Close'].isnull().all():
                        close_price = float(row_df['Close'].iloc[0])
                        break
                search_date -= timedelta(days=1)
            return close_price
        price_company = find_price(hist_company, this_date)
        price_index = find_price(hist_index, this_date)
        ws_dst[f"A{row}"] = this_date.strftime("%Y-%m-%d")
        ws_dst[f"B{row}"] = price_company if price_company is not None else "Veri yok"
        ws_dst[f"C{row}"] = price_index if price_index is not None else "Veri yok"
        prev_date = this_date

def create_final2_file_for_ticker(ticker):
    final_folder = s3_path("Final")
    final2_folder = s3_path("Final2")
    template_path = s3_path("Companies1/donusturucu.xlsx")
    src_key = s3_path(f"Final/{ticker}.xlsx")
    if not s3_exists(src_key):
        return
    if not s3_exists(template_path):
        return
    try:
        wb_src = openpyxl.load_workbook(BytesIO(s3_read_bytes(src_key)), data_only=True)
        ws_src = wb_src.active
        dst_key = s3_path(f"Final2/{ticker}.xlsx")
        template_bytes = s3_read_bytes(template_path)
        s3_write_bytes(dst_key, template_bytes)
        wb_dst = openpyxl.load_workbook(BytesIO(template_bytes), data_only=False)
        ws_dst = wb_dst.active
        from openpyxl.utils import column_index_from_string
        max_col = column_index_from_string("BG")
        max_row = 36
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                src_val = ws_src.cell(row=r, column=c).value
                if src_val is not None:
                    ws_dst.cell(row=r, column=c).value = src_val
        ws_dst["B40"].value = ticker
        ws_dst["C40"].value = "^GSPC"
        fill_dates_and_prices_in_ws(ws_dst)
        yf_ticker = yf.Ticker(ticker)
        try:
            ws_dst["E41"].value = yf_ticker.info.get("sector", "")
        except Exception as e:
            ws_dst["E41"].value = ""
        try:
            ws_dst["F41"].value = yf_ticker.info.get("industry", "")
        except Exception as e:
            ws_dst["F41"].value = ""
        try:
            ws_dst["G41"].value = yf_ticker.info.get("fullTimeEmployees", "")
        except Exception as e:
            ws_dst["G41"].value = ""
        try:
            summary = yf_ticker.info.get("longBusinessSummary", yf_ticker.info.get("summary", ""))
            ws_dst["H41"].value = summary
        except Exception as e:
            ws_dst["H41"].value = ""
        try:
            ws_dst["E45"].value = yf_ticker.info.get("beta", "")
        except Exception as e:
            ws_dst["E45"].value = ""
        try:
            tnx_ticker = yf.Ticker("^TNX")
            tnx_yield = tnx_ticker.info.get("regularMarketPrice", "")
            ws_dst["F45"].value = tnx_yield
        except Exception as e:
            ws_dst["F45"].value = ""
        try:
            cal = yf_ticker.calendar
            earning_date = ""
            if isinstance(cal, pd.DataFrame):
                if not cal.empty and "Earnings Date" in cal.index:
                    earning_date = cal.loc["Earnings Date"][0]
            elif isinstance(cal, dict):
                earning_date = cal.get("Earnings Date", [None])[0]
            ws_dst["I41"].value = str(earning_date) if earning_date else ""
        except Exception as e:
            ws_dst["I41"].value = ""
        try:
            with open("excel python donusum.txt", "r", encoding="utf-8") as f:
                formul_code = f.read()
            formul_ns = {}
            exec(formul_code, formul_ns)
            formul_ns["hesapla_tum_formuller"](ws_dst)
        except Exception as e:
            pass
        buffer = BytesIO()
        wb_dst.save(buffer)
        buffer.seek(0)
        s3_write_bytes(dst_key, buffer.read())
    except Exception as e:
        pass

def get_data_from_excel(filepath, range_tuple):
    wb = openpyxl.load_workbook(BytesIO(s3_read_bytes(filepath)), data_only=True)
    ws = wb.active
    start_col, start_row = range_tuple[0][0], int(range_tuple[0][1:])
    end_col, end_row = range_tuple[1][0], int(range_tuple[1][1:])
    data = []
    for row in ws.iter_rows(min_row=start_row, max_row=end_row,
                            min_col=ord(start_col)-64, max_col=ord(end_col)-64):
        values = [cell.value for cell in row]
        data.append(values)
    return data

def insert_company_info_to_db(cursor, ticker, sector, industry, employees, earnings_date, summary, radar=None, market_cap=None):
    sql = """
        INSERT INTO company_info (ticker, sector, industry, employees, earnings_date, summary, radar, market_cap)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (ticker) DO UPDATE
        SET sector=EXCLUDED.sector,
            industry=EXCLUDED.industry,
            employees=EXCLUDED.employees,
            earnings_date=EXCLUDED.earnings_date,
            summary=EXCLUDED.summary,
            radar=EXCLUDED.radar,
            market_cap=EXCLUDED.market_cap
    """
    cursor.execute(sql, (ticker, sector, industry, employees, earnings_date, summary, radar, market_cap))

def insert_data_to_db(cursor, ticker, data):
    headers = data[0]
    for row in data[1:]:
        metric = row[0]
        for i in range(1, len(headers)):
            period = headers[i]
            value = row[i]
            if period is None or period == 0 or str(period).strip() == "":
                continue
            if value is None or str(value).strip() == "" or str(value).startswith("#VALUE"):
                continue
            sql = """
            INSERT INTO excel_metrics (ticker, metric, period, value)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (ticker, metric, period)
            DO UPDATE SET value = EXCLUDED.value
            """
            cursor.execute(sql, (ticker, metric, period, str(value) if value is not None else None))

def upload_to_db(ticker):
    db_user = os.getenv("DB_USER")
    db_pass = os.getenv("DB_PASS")
    db_host = os.getenv("DB_HOST")
    db_port = os.getenv("DB_PORT", "5432")
    db_name = os.getenv("DB_NAME")
    conn = psycopg2.connect(
        host=db_host,
        port=db_port,
        user=db_user,
        password=db_pass,
        dbname=db_name
    )
    cursor = conn.cursor()
    fpath = s3_path(f"Final2/{ticker}.xlsx")
    if not s3_exists(fpath):
        return
    try:
        EXCEL_RANGE = ('A191', 'O202')
        data = get_data_from_excel(fpath, EXCEL_RANGE)
        insert_data_to_db(cursor, ticker, data)
        conn.commit()
        wb = openpyxl.load_workbook(BytesIO(s3_read_bytes(fpath)), data_only=True)
        ws = wb.active
        sector = ws["B204"].value
        industry = ws["B205"].value
        employees = ws["B206"].value
        earnings_date = ws["B207"].value
        summary = ws["B208"].value
        radar = ws["B209"].value
        market_cap = ws["B210"].value
        insert_company_info_to_db(cursor, ticker, sector, industry, employees, earnings_date, summary,
                                  int(radar) if radar is not None else None,
                                  int(market_cap) if market_cap is not None else None)
        conn.commit()
    except Exception as e:
        pass
    cursor.close()
    conn.close()

if __name__ == "__main__":
    main()
