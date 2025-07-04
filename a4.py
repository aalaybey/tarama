import sys
import argparse
import os
import json
import requests
import shutil
from datetime import datetime, timedelta, UTC
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
import re
import portalocker
import time
import threading
import boto3
from io import BytesIO, StringIO

consecutive_error_count = 0

REQUEST_COUNT = 0
REQUEST_LOCK = threading.Lock()

AWS_BUCKET = "alaybey"
S3_PREFIX = "s3/"
AWS_REGION = os.getenv("AWS_REGION")
AWS_ACCESS_KEY_ID = os.getenv("AWS_ACCESS")
AWS_SECRET_ACCESS_KEY = os.getenv("AWS_SECRET")

s3 = boto3.client(
    "s3",
    region_name=AWS_REGION,
    aws_access_key_id=AWS_ACCESS_KEY_ID,
    aws_secret_access_key=AWS_SECRET_ACCESS_KEY,
)

def s3_path(key):  # Her path için kullan
    return S3_PREFIX + key if not key.startswith(S3_PREFIX) else key

def s3_exists(key):
    try:
        s3.head_object(Bucket=AWS_BUCKET, Key=s3_path(key))
        return True
    except:
        return False

def s3_read_text(key):
    obj = s3.get_object(Bucket=AWS_BUCKET, Key=s3_path(key))
    return obj["Body"].read().decode("utf-8")

def s3_read_bytes(key):
    obj = s3.get_object(Bucket=AWS_BUCKET, Key=s3_path(key))
    return obj["Body"].read()

def s3_write_text(key, text):
    s3.put_object(Bucket=AWS_BUCKET, Key=s3_path(key), Body=text.encode("utf-8"))

def s3_write_bytes(key, b):
    s3.put_object(Bucket=AWS_BUCKET, Key=s3_path(key), Body=b)

def s3_delete(key):
    s3.delete_object(Bucket=AWS_BUCKET, Key=s3_path(key))

def s3_list_dir(prefix):
    paginator = s3.get_paginator("list_objects_v2")
    result = paginator.paginate(Bucket=AWS_BUCKET, Prefix=s3_path(prefix))
    return [content['Key'] for page in result for content in page.get('Contents', [])]

def incr_request_and_sleep():
    global REQUEST_COUNT
    with REQUEST_LOCK:
        REQUEST_COUNT += 1
        if REQUEST_COUNT % 10 == 0:
            print(f"🕒 {REQUEST_COUNT} request atıldı, 2 sn bekleniyor...")
            time.sleep(0.5)

def inc_error_and_kill_if_limit(limit=50):
    global consecutive_error_count
    consecutive_error_count += 1
    if consecutive_error_count >= limit:
        print("LOG: Script kill oldu! (üst üste hata limiti aşıldı)")
        print(f"❌ ÜST ÜSTE {limit} HATA! Script kill ediliyor.")
        sys.exit(1)

parser = argparse.ArgumentParser()
parser.add_argument("--proxy", type=str, required=False, default=None)
parser.add_argument("--script_id", type=int, required=False, default=1)
parser.add_argument("--user_agent", type=str, required=False, default="a.alaybey@gmail.com")
args, unknown = parser.parse_known_args()

PROXY = args.proxy
SCRIPT_ID = args.script_id
USER_AGENT = args.user_agent

if PROXY:
    print(f"[Script {SCRIPT_ID}] Proxy ile başlatıldı: {PROXY}")

REQUESTS_PROXIES = {"http": PROXY, "https": PROXY} if PROXY else None

PAUSE_FLAG = False
STOP_FLAG = False
PROGRESS_BAR = None

HEADERS = {"User-Agent": USER_AGENT}
REPORT_DIR = f"Companies{SCRIPT_ID}"
CHECKPOINT_FILE = f"a{SCRIPT_ID}_checkpoint.txt"

def save_checkpoint(symbol):
    s3_write_text(CHECKPOINT_FILE, symbol)
    print(f"LOG: Checkpoint kaydedildi: {symbol}")

def load_checkpoint():
    if s3_exists(CHECKPOINT_FILE):
        return s3_read_text(CHECKPOINT_FILE).strip()
    return None

def clear_checkpoint():
    if s3_exists(CHECKPOINT_FILE):
        s3_delete(CHECKPOINT_FILE)

def load_companies():
    path = f"a{SCRIPT_ID}.txt"
    if not s3_exists(path):
        print(f"❌ {path} bulunamadı.")
        inc_error_and_kill_if_limit()
        return {}
    companies = {}
    try:
        content = s3_read_text(path)
        for line in content.splitlines():
            line = line.strip()
            if not line:
                continue
            if ',' not in line:
                continue
            ticker, cik = line.split(",", 1)
            ticker = ticker.strip().upper()
            cik = cik.strip().zfill(10)
            if ticker and cik:
                companies[ticker] = cik
    except Exception as e:
        print(f"❌ {path} okunamadı: {e}")
        inc_error_and_kill_if_limit()
        return {}
    if not companies:
        print("❌ Hiç şirket bulunamadı (txt dosyası boş mu?)")
        inc_error_and_kill_if_limit()
    return companies

def download_file(url, s3key):
    backoff_count = 0
    while True:
        try:
            incr_request_and_sleep()
            resp = requests.get(url, headers=HEADERS, proxies=REQUESTS_PROXIES)
            if resp.status_code in [429, 403]:
                backoff_count += 1
                print(f"⏳ Rate-limit algılandı! {backoff_count}. kez 2 dakika bekleniyor... [download_file] {url}")
                if backoff_count >= 3:
                    print("❌ 3 kez üst üste backoff, script kill ediliyor!")
                    sys.exit(1)
                time.sleep(120)
                continue
            if resp.status_code >= 400:
                print(f"📥 Dosya indirme hatası: {url} => {resp.status_code} {resp.reason}")
                break
            s3_write_bytes(s3key, resp.content)
            break
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout):
            backoff_count += 1
            print(f"🌐 Bağlantı hatası: {url}")
            if backoff_count >= 3:
                print("❌ 3 kez üst üste backoff, script kill ediliyor!")
                sys.exit(1)
            time.sleep(120)
            continue
        except Exception as e:
            print(f"📥 Dosya indirme hatası: {url} => {e}")
            inc_error_and_kill_if_limit()
            break

def download_xlsx(index_url, folder_path, file_name, is_10k=False):
    print(f"🔎 XLSX aranıyor: {file_name} [{index_url}]")
    backoff_count = 0
    while True:
        try:
            incr_request_and_sleep()
            resp = requests.get(index_url, headers=HEADERS, proxies=REQUESTS_PROXIES)
            if resp.status_code in [429, 403]:
                backoff_count += 1
                print(f"⏳ Rate-limit algılandı! {backoff_count}. kez 2 dakika bekleniyor... [download_xlsx] {index_url}")
                if backoff_count >= 3:
                    print("❌ 3 kez üst üste backoff, script kill ediliyor!")
                    sys.exit(1)
                time.sleep(120)
                continue
            if resp.status_code >= 400:
                print(f"❌ Index sayfası hatası: {index_url} => {resp.status_code} {resp.reason}")
                return None, None, None, None
            break
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout):
            backoff_count += 1
            print(f"🌐 Bağlantı hatası: {index_url}")
            if backoff_count >= 3:
                print("❌ 3 kez üst üste backoff, script kill ediliyor!")
                sys.exit(1)
            time.sleep(120)
            continue
        except Exception as e:
            print(f"❌ Index sayfası alınamadı: {e}")
            inc_error_and_kill_if_limit()
            return None, None, None, None

    soup = BeautifulSoup(resp.text, "html.parser")
    xlsx_url = None
    for link in soup.find_all("a"):
        href = link.get("href", "")
        if href.lower().endswith(".xlsx"):
            if href.startswith("/"):
                xlsx_url = "https://www.sec.gov" + href
            else:
                base = index_url.rsplit("/", 1)[0]
                xlsx_url = base + "/" + href
            break

    if not xlsx_url:
        return None, None, None, None

    temp_path = f"{folder_path}/temp.xlsx"
    if s3_exists(temp_path):
        try:
            s3_delete(temp_path)
            print(f"ℹ️ Eski temp.xlsx dosyası silindi: {temp_path}")
        except Exception as e:
            print(f"⚠️ Eski temp dosyası silinemedi: {e}")

    download_file(xlsx_url, temp_path)

    if not s3_exists(temp_path):
        print(f"⚠️ Dosya temp.xlsx oluşturulamadı: {temp_path}")
        inc_error_and_kill_if_limit()
        return None, None, None, None

    try:
        b = s3_read_bytes(temp_path)
        wb = openpyxl.load_workbook(BytesIO(b), data_only=True)
        year, quarter = None, None
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and "Document Fiscal Year Focus" in cell.value:
                        idx = row.index(cell)
                        if idx + 1 < len(row):
                            year_cell = row[idx + 1]
                            if year_cell.value is not None:
                                year = str(year_cell.value).strip()
            if is_10k:
                quarter = "Q4"
        if not is_10k:
            for sheet in wb.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and "Document Fiscal Period Focus" in cell.value:
                            idx = row.index(cell)
                            if idx + 1 < len(row):
                                qval = row[idx + 1].value
                                if qval is not None:
                                    qval = str(qval).strip().upper()
                                    if qval.startswith("Q"):
                                        quarter = qval
                                    else:
                                        quarter = "Q" + qval
        base_ticker = file_name.split("-")[0]
        if not year or not quarter:
            parts = file_name.split("-")
            if len(parts) >= 3:
                year = parts[1]
                quarter = parts[2]
        symbol = base_ticker
        new_base = f"{symbol}-{year}-{quarter}"
        final_name = new_base + ".xlsx"
        final_path = f"{folder_path}/{final_name}"

        if s3_exists(final_path):
            print(f"ℹ️ Zaten var, atlandı: {final_path}")
            s3_delete(temp_path)
        else:
            b = s3_read_bytes(temp_path)
            s3_write_bytes(final_path, b)
            s3_delete(temp_path)
            print(f"✅ Kaydedildi: {final_path}")

        return final_path, symbol, year, quarter

    except Exception as e:
        print(f"⚠️ XLSX işleme hatası: {e}")
        fallback = f"{folder_path}/{file_name}.xlsx"
        b = s3_read_bytes(temp_path)
        s3_write_bytes(fallback, b)
        s3_delete(temp_path)
        print(f"⚠️ Geçici adla kaydedildi: {fallback}")
        inc_error_and_kill_if_limit()
        parts = file_name.split("-")
        if len(parts) >= 3:
            symbol = parts[0]
            year = parts[1]
            quarter = parts[2]
            return fallback, symbol, year, quarter
        return fallback, None, None, None

def get_sheet_title(sheet):
    try:
        top_left = sheet["A1"].value
        if isinstance(top_left, str):
            return top_left.strip().lower()
        else:
            return ""
    except:
        return ""

def detect_multiplier_from_title(title):
    if not title:
        return 1.0
    words = title.split()
    if not words:
        return 1.0
    last = re.sub(r"[^\w]", "", words[-1]).lower()
    if last == "millions":
        return 1_000_000.0
    elif last == "thousands":
        return 1_000.0
    else:
        return 1.0

def find_metric_value(sheet, keyword_groups, multiplier=1.0, exclude_term=None, reverse=False, quarter=None, row_start=None):
    rows = list(sheet.iter_rows())
    if reverse:
        rows = rows[::-1]
    if row_start is not None:
        rows = rows[row_start:]
    col_idx = None
    if quarter is not None:
        quarter_map = {"Q1": "3 month", "Q2": "6 month", "Q3": "9 month", "Q4": "12 month"}
        target_phrase = quarter_map.get(quarter.upper())
        if target_phrase:
            for i in range(min(2, len(rows))):
                for j, cell in enumerate(rows[i]):
                    if cell.value and target_phrase.lower() in str(cell.value).lower():
                        col_idx = j
                        break
                if col_idx is not None:
                    break
    for group in keyword_groups:
        for row in rows:
            row_text = " ".join(str(c.value).lower() for c in row if c.value is not None)
            if exclude_term and exclude_term.lower() in row_text:
                continue
            if all(k.lower() in row_text for k in group):
                label_idx = None
                for idx, c in enumerate(row):
                    if isinstance(c.value, str):
                        cell_text = c.value.lower()
                        if all(k.lower() in cell_text for k in group):
                            label_idx = idx
                            break
                if label_idx is not None:
                    if col_idx is not None:
                        c = row[col_idx]
                        if isinstance(c.value, (int, float)):
                            return c.value * multiplier
                    for c in row[label_idx + 1:]:
                        if isinstance(c.value, (int, float)):
                            return c.value * multiplier
                    for c in row:
                        if isinstance(c.value, (int, float)):
                            return c.value * multiplier
    return None

def s3_load_workbook(key):
    b = s3_read_bytes(key)
    return openpyxl.load_workbook(BytesIO(b), data_only=True)

def s3_write_excel(df, key):
    out = BytesIO()
    df.to_excel(out, index=False)
    out.seek(0)
    s3_write_bytes(key, out.read())

def extract_metrics(file_path, symbol, year, quarter):
    print(f"📊 Veriler çıkarılıyor: {symbol} - {year} {quarter}")

    try:
        wb = s3_load_workbook(file_path)
    except Exception as e:
        print(f"⚠️ Excel açılamadı: {file_path} => {e}")
        inc_error_and_kill_if_limit()
        return

    cash_flow_metrics = {
        "Depreciation and Amortization": [["depreciation", "amortization"]],
        "Amortization": [["amortization"]],
        "Depreciation": [["depreciation"]],
        "Cash From Operations": [["cash", "operat"]],
        "PPE Purchase": [["property", "equipment", "purchas"], ["property", "equipment", "add"], ["property", "equipment", "payment"], ["property", "equipment", "acqui"], ["property", "add"], ["property", "purchas"], ["property", "acqui"]],
        "PPE Sale": [["sale", "property"], ["sale", "fixed"], ["disposal", "property"], ["disposal", "fixed"], ["sale", "tangible"], ["disposal", "tangible"]],
        "Cash Business Acquisitions": [["business"]],
        "Dividends": [["payment", "dividend"]],
        "Cash Taxes": [["paid", "tax"], ["cash", "tax"], ["tax"]],
        "Cash Interest": [["paid", "interest"], ["cash", "interest"], ["interest"]],
    }

    income_metrics = {
        "Sales": [["revenue"], ["sale"]],
        "Cost of Sales": [["cost of"]],
        "Gross Profit": [["gross"], ["margin"], ["gross", "profit"]],
        "EBIT": [["operating", "income"], ["operating", "loss"], ["income", "operations"], ["loss", "operations"]],
        "Net Income": [["net income"], ["net loss"], ["net", "income"], ["net", "loss"]],
        "Interest Income": [["interest", "income"], ["other", "net"]],
        "Interest Expense": [["interest", "expense"]],
        "EPS": [["basic", "net", "income"], ["basic", "net", "loss"], ["basic", "dollar"], ["basic", "usd"]],
        "Shares Outstanding": [["basic"]]
    }

    balance_metrics = {
        "Total Assets": [["total assets"]],
        "Total Equity": [["total", "equity"]],
        "Noncontrolling interest": [["non", "control", "interest"]],
        "Cash and cash equivalents": [["cash", "equivalents"]],
        "Inventories": [["inventor"]],
        "Accounts receivable": [["accounts", "receivable"]],
        "Prepaid expenses": [["prepaid", "expenses"]],
        "PPE": [["property", "equipment"]],
        "Intangible assets": [["intangible", "asset"]],
        "Operating lease assets": [["leas", "operati", "asset"]],
        "Digital assets": [["digital", "asset"]],
        "Right-of-use assets": [["use", "asset"]],
        "Accounts payable": [["account", "payable"]],
        "Operating lease liabilities": [["leas", "operati", "liab"]],
        "Accrued liabilities": [["accrue", "liabilit"], ["accrue", "expense"]],
        "Deferred revenue": [["defer", "revenue"], ["unearn", "revenue"]]
    }

    extracted = {}

    cash_flow_sheet = None
    cash_flow_multiplier = 1.0
    for sheet in wb.worksheets:
        title = get_sheet_title(sheet)
        if "cash flow" in title:
            cash_flow_sheet = sheet
            cash_flow_multiplier = detect_multiplier_from_title(title)
            break

    if cash_flow_sheet is not None:
        for metric, kw_groups in cash_flow_metrics.items():
            val = find_metric_value(
                cash_flow_sheet, kw_groups, multiplier=cash_flow_multiplier, exclude_term=None, reverse=False, quarter=quarter
            )
            extracted[metric] = val if val is not None else 0.0
    else:
        for metric in cash_flow_metrics.keys():
            extracted[metric] = 0.0

    income_sheet = None
    income_general_multiplier = 1.0
    income_share_multiplier = None

    for sheet in wb.worksheets:
        title = get_sheet_title(sheet)
        if "statements of operations" in title:
            income_sheet = sheet
            income_general_multiplier = detect_multiplier_from_title(title)
            low = title.lower()
            if "shares in millions" in low:
                income_share_multiplier = 1_000_000.0
            elif "shares in thousands" in low:
                income_share_multiplier = 1_000.0
            else:
                income_share_multiplier = 1.0
            break

    if income_sheet is None:
        for sheet in wb.worksheets:
            title = get_sheet_title(sheet)
            if "income" in title and "statement" in title:
                income_sheet = sheet
                income_general_multiplier = detect_multiplier_from_title(title)
                low = title.lower()
                if "shares in millions" in low:
                    income_share_multiplier = 1_000_000.0
                elif "shares in thousands" in low:
                    income_share_multiplier = 1_000.0
                else:
                    income_share_multiplier = 1.0
                break

    if income_sheet is not None:
        for metric, kw_groups in income_metrics.items():
            if metric == "Shares Outstanding":
                multiplier = income_share_multiplier or 1.0
                value = find_metric_value(income_sheet, kw_groups, multiplier=multiplier, exclude_term=None, reverse=True, quarter=quarter)
            else:
                value = find_metric_value(income_sheet, kw_groups, multiplier=income_general_multiplier, exclude_term=None, reverse=False, quarter=quarter)
            extracted[metric] = value if value is not None else 0.0
    else:
        for metric in income_metrics.keys():
            extracted[metric] = 0.0

    balance_sheet = None
    balance_multiplier = 1.0

    for sheet in wb.worksheets:
        title = get_sheet_title(sheet)
        if "balance sheet" in title and "parenthetical" not in title:
            balance_sheet = sheet
            balance_multiplier = detect_multiplier_from_title(title)
            break

    if balance_sheet is not None:
        for metric, kw_groups in balance_metrics.items():
            value = find_metric_value(balance_sheet, kw_groups, multiplier=balance_multiplier, exclude_term=None, reverse=False, quarter=quarter)
            extracted[metric] = value if value is not None else 0.0
    else:
        for metric in balance_metrics.keys():
            extracted[metric] = 0.0

    s3_out_dir = "Final"
    out_path = f"{s3_out_dir}/{symbol}.xlsx"

    all_metrics = list(cash_flow_metrics.keys()) + list(income_metrics.keys()) + list(balance_metrics.keys())

    if s3_exists(out_path):
        df_existing = pd.read_excel(BytesIO(s3_read_bytes(out_path)))
        if "Metric" not in df_existing.columns:
            df_existing.insert(0, "Metric", all_metrics)
        df = df_existing.set_index("Metric")
    else:
        df = pd.DataFrame(index=all_metrics)

    col_name = f"{year} {quarter}"
    col_values = [extracted.get(m, 0.0) for m in all_metrics]
    df[col_name] = col_values

    df_to_save = df.reset_index().rename(columns={"index": "Metric"})
    s3_write_excel(df_to_save, out_path)

def full_process(days=None):
    global PAUSE_FLAG, STOP_FLAG, PROGRESS_BAR
    # Klasör oluşturma yok, doğrudan S3'e yazılıyor
    companies = load_companies()
    if not companies:
        print("Hiç şirket bulunamadı!")
        inc_error_and_kill_if_limit()
        return

    checkpoint_symbol = load_checkpoint()
    skip_mode = bool(checkpoint_symbol)
    symbols = list(companies.keys())
    total = len(symbols)
    processed = 0
    found_checkpoint = False

    if PROGRESS_BAR is not None:
        PROGRESS_BAR['maximum'] = total
        PROGRESS_BAR['value'] = 0

    for symbol in symbols:
        if STOP_FLAG:
            print("🔴 İşlem kullanıcı tarafından durduruldu.")
            break
        while PAUSE_FLAG:
            if STOP_FLAG:
                break
        if STOP_FLAG:
            print("🔴 İşlem kullanıcı tarafından durduruldu.")
            break

        if skip_mode and not found_checkpoint:
            if symbol == checkpoint_symbol:
                found_checkpoint = True
            else:
                continue

        cik = companies[symbol]
        backoff_count = 0
        while True:
            try:
                url = f"https://data.sec.gov/submissions/CIK{cik}.json"
                incr_request_and_sleep()
                r = requests.get(url, headers=HEADERS, proxies=REQUESTS_PROXIES)
                if r.status_code in [429, 403]:
                    backoff_count += 1
                    print(f"⏳ Rate-limit algılandı! {backoff_count}. kez 2 dakika bekleniyor... [full_process] {url}")
                    if backoff_count >= 3:
                        print("❌ 3 kez üst üste backoff, script kill ediliyor!")
                        sys.exit(1)
                    time.sleep(120)
                    continue
                if r.status_code >= 400:
                    print(f"⚠️ {symbol}: {cik} - API hatası veya veri yok. {r.status_code} {r.reason}")
                    inc_error_and_kill_if_limit()
                    break
                data = r.json().get("filings", {}).get("recent", {})
                break
            except (requests.exceptions.ConnectionError, requests.exceptions.Timeout):
                backoff_count += 1
                print(f"🌐 Bağlantı hatası: {url}")
                if backoff_count >= 3:
                    print("❌ 3 kez üst üste backoff, script kill ediliyor!")
                    sys.exit(1)
                time.sleep(120)
                continue
            except Exception as e:
                print(f"⚠️ {symbol}: {cik} - Veri çekilemedi: {e}")
                inc_error_and_kill_if_limit()
                break

        forms = data.get("form", [])
        accessions = data.get("accessionNumber", [])
        report_dates = data.get("reportDate", [])

        for i, ftype in enumerate(forms):
            if STOP_FLAG:
                break
            while PAUSE_FLAG:
                if STOP_FLAG:
                    break
            if STOP_FLAG:
                break

            if ftype not in ["10-Q", "10-K"]:
                continue
            try:
                rpt_date = datetime.strptime(report_dates[i], "%Y-%m-%d")
            except:
                continue
            if days is not None:
                filter_date = (datetime.now(UTC) - timedelta(days=days)).date()
                if rpt_date.date() < filter_date:
                    continue
            acc = accessions[i].replace("-", "")
            year = str(rpt_date.year)
            month = rpt_date.month
            quarter_idx = (month - 1) // 3 + 1
            is_10k = (ftype == "10-K")
            quarter_label = f"Q{quarter_idx}" if not is_10k else "Q4"
            folder = f"{REPORT_DIR}/{symbol}"
            index_url = f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{acc}/index.html"
            fname = f"{symbol}-{year}-{quarter_label}"
            file_path, sym, y, q = download_xlsx(index_url, folder, fname, is_10k)
            if file_path and sym and y and q:
                extract_metrics(file_path, sym, y, q)
                global consecutive_error_count
                consecutive_error_count = 0
            else:
                pass

        save_checkpoint(symbol)

        processed += 1
        if PROGRESS_BAR is not None:
            PROGRESS_BAR.step(1)
            PROGRESS_BAR.update()

    if not STOP_FLAG:
        print("✅ Tüm işlemler tamamlandı.")
        clear_checkpoint()
    PAUSE_FLAG = False
    STOP_FLAG = False

if __name__ == "__main__":
    full_process(days=1)
